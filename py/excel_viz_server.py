from mcp.server.fastmcp import FastMCP
import pandas as pd
import polars as pl
import os
import json
import base64
import io
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib
import numpy as np

matplotlib.use('Agg')  # Nécessaire pour le mode non-interactif

# Initialiser le serveur MCP
mcp = FastMCP("excel-viz-server")

# Chemin par défaut pour les fichiers Excel
DEFAULT_EXCEL_DIR = os.path.expanduser("~\\Documents")

# Taille limite pour choisir entre Pandas et Polars (en Mo)
LARGE_FILE_THRESHOLD = 50  # MB


def is_large_file(file_path):
    """Détermine si un fichier est considéré comme volumineux"""
    return os.path.getsize(file_path) / (1024 * 1024) > LARGE_FILE_THRESHOLD


@mcp.tool()
def read_excel(file_path: str, sheet_name: str = None) -> str:
    """
    Lit un fichier Excel et retourne son contenu sous forme de texte.

    Args:
        file_path: Chemin du fichier Excel (peut être relatif au répertoire par défaut)
        sheet_name: Nom de la feuille à lire (si None, lit la première feuille)

    Returns:
        Contenu du fichier Excel sous forme tabulaire
    """
    try:
        # Gérer les chemins relatifs
        if not os.path.isabs(file_path):
            file_path = os.path.join(DEFAULT_EXCEL_DIR, file_path)

        # Vérifier que le fichier existe
        if not os.path.exists(file_path):
            return f"Erreur: Le fichier '{file_path}' n'existe pas."

        # Déterminer si le fichier est volumineux
        large_file = is_large_file(file_path)

        # Lire le fichier Excel
        if large_file:
            # Utiliser Polars pour les fichiers volumineux
            # Note: Polars n'a pas de support natif pour Excel, on utilise pandas pour charger
            # puis on convertit
            if sheet_name:
                df_pd = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df_pd = pd.read_excel(file_path)

            # Convertir en DataFrame Polars
            df = pl.from_pandas(df_pd)

            # Extraire un échantillon pour l'affichage si trop grand
            if len(df) > 1000:
                result_df = df.head(1000)
                is_sample = True
            else:
                result_df = df
                is_sample = False

            # Convertir en chaîne formatée
            result = result_df.to_pandas().to_string(index=False)

            # Ajouter des informations sur le fichier
            header = f"Fichier: {os.path.basename(file_path)} (traité avec Polars pour performance optimale)\n"
            header += f"Feuille: {sheet_name or 'par défaut'}\n"
            header += f"Dimensions totales: {len(df)} lignes × {len(df.columns)} colonnes\n"
            if is_sample:
                header += f"Affichage limité aux 1000 premières lignes\n"
            header += "\n"

            return header + result
        else:
            # Utiliser Pandas pour les fichiers standard
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)

            # Convertir en chaîne formatée
            result = df.to_string(index=False)

            # Ajouter des informations sur le fichier
            header = f"Fichier: {os.path.basename(file_path)}\n"
            header += f"Feuille: {sheet_name or 'par défaut'}\n"
            header += f"Dimensions: {df.shape[0]} lignes × {df.shape[1]} colonnes\n\n"

            return header + result
    except Exception as e:
        return f"Erreur lors de la lecture du fichier Excel: {str(e)}"


@mcp.tool()
def get_excel_sheets(file_path: str) -> str:
    """
    Liste toutes les feuilles dans un fichier Excel.

    Args:
        file_path: Chemin du fichier Excel (peut être relatif au répertoire par défaut)

    Returns:
        Liste des noms de feuilles dans le fichier Excel
    """
    try:
        # Gérer les chemins relatifs
        if not os.path.isabs(file_path):
            file_path = os.path.join(DEFAULT_EXCEL_DIR, file_path)

        # Vérifier que le fichier existe
        if not os.path.exists(file_path):
            return f"Erreur: Le fichier '{file_path}' n'existe pas."

        # Lire les noms de feuilles
        xls = pd.ExcelFile(file_path)
        sheets = xls.sheet_names

        # Formater la réponse
        result = f"Feuilles dans '{os.path.basename(file_path)}':\n\n"
        for i, sheet in enumerate(sheets, 1):
            result += f"{i}. {sheet}\n"

        return result
    except Exception as e:
        return f"Erreur lors de la lecture des feuilles: {str(e)}"


@mcp.tool()
def excel_summary(file_path: str, sheet_name: str = None) -> str:
    """
    Génère un résumé statistique des données dans un fichier Excel.

    Args:
        file_path: Chemin du fichier Excel (peut être relatif au répertoire par défaut)
        sheet_name: Nom de la feuille à analyser (si None, utilise la première feuille)

    Returns:
        Résumé statistique des données
    """
    try:
        # Gérer les chemins relatifs
        if not os.path.isabs(file_path):
            file_path = os.path.join(DEFAULT_EXCEL_DIR, file_path)

        # Vérifier que le fichier existe
        if not os.path.exists(file_path):
            return f"Erreur: Le fichier '{file_path}' n'existe pas."

        # Déterminer si le fichier est volumineux
        large_file = is_large_file(file_path)

        if large_file:
            # Utiliser Polars pour les fichiers volumineux
            if sheet_name:
                df_pd = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df_pd = pd.read_excel(file_path)

            # Convertir en DataFrame Polars
            df_pl = pl.from_pandas(df_pd)

            # Créer un résumé statistique à la main (Polars n'a pas d'équivalent direct à describe())
            num_cols = []
            for col in df_pl.columns:
                if df_pl[col].dtype in [pl.Float32, pl.Float64, pl.Int32, pl.Int64]:
                    num_cols.append(col)

            summary_data = {}
            for col in num_cols:
                summary_data[col] = {
                    'count': df_pl[col].count(),
                    'mean': df_pl[col].mean(),
                    'std': df_pl[col].std(),
                    'min': df_pl[col].min(),
                    '25%': df_pl[col].quantile(0.25),
                    '50%': df_pl[col].median(),
                    '75%': df_pl[col].quantile(0.75),
                    'max': df_pl[col].max()
                }

            # Convertir les statistiques en DataFrame pandas pour un affichage cohérent
            summary_df = pd.DataFrame(summary_data)
            summary = summary_df.to_string()

            # Informations supplémentaires
            info = f"Fichier: {os.path.basename(file_path)} (traité avec Polars pour performance optimale)\n"
            info += f"Feuille: {sheet_name or 'par défaut'}\n"
            info += f"Dimensions: {len(df_pl)} lignes × {len(df_pl.columns)} colonnes\n"
            info += f"Colonnes: {', '.join(df_pl.columns)}\n\n"

            return info + summary
        else:
            # Utiliser pandas pour les fichiers standards
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)

            # Générer résumé statistique
            summary = df.describe(include='all').fillna("N/A").to_string()

            # Informations supplémentaires
            info = f"Fichier: {os.path.basename(file_path)}\n"
            info += f"Feuille: {sheet_name or 'par défaut'}\n"
            info += f"Dimensions: {df.shape[0]} lignes × {df.shape[1]} colonnes\n"
            info += f"Colonnes: {', '.join(df.columns.tolist())}\n\n"

            return info + summary
    except Exception as e:
        return f"Erreur lors de la génération du résumé: {str(e)}"


@mcp.tool()
def excel_query(file_path: str, query: str, sheet_name: str = None) -> str:
    """
    Exécute une requête simple sur les données Excel.

    Args:
        file_path: Chemin du fichier Excel (peut être relatif au répertoire par défaut)
        query: Requête sous forme de texte (ex: "colonne > 100")
        sheet_name: Nom de la feuille à analyser (si None, utilise la première feuille)

    Returns:
        Résultat de la requête
    """
    try:
        # Gérer les chemins relatifs
        if not os.path.isabs(file_path):
            file_path = os.path.join(DEFAULT_EXCEL_DIR, file_path)

        # Vérifier que le fichier existe
        if not os.path.exists(file_path):
            return f"Erreur: Le fichier '{file_path}' n'existe pas."

        # Déterminer si le fichier est volumineux
        large_file = is_large_file(file_path)

        if large_file:
            # Utiliser Polars pour les fichiers volumineux
            if sheet_name:
                df_pd = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df_pd = pd.read_excel(file_path)

            # Convertir en DataFrame Polars
            df = pl.from_pandas(df_pd)

            # Adapter la requête au format Polars
            # Note: ceci est une implémentation simplifiée, certaines requêtes complexes
            # peuvent nécessiter une conversion plus sophistiquée
            try:
                # Convertir le format de requête pandas en polars
                # Cette conversion est simplifiée et ne gère pas tous les cas
                query_parts = query.split()
                if len(query_parts) >= 3:
                    col_name = query_parts[0]
                    operator = query_parts[1]
                    value = ' '.join(query_parts[2:])

                    # Essayer de convertir la valeur en nombre si possible
                    try:
                        if '.' in value:
                            value = float(value)
                        else:
                            value = int(value)
                    except ValueError:
                        # Si c'est une chaîne de caractères, supprimer les guillemets
                        if value.startswith("'") and value.endswith("'"):
                            value = value[1:-1]
                        elif value.startswith('"') and value.endswith('"'):
                            value = value[1:-1]

                    # Construire l'expression Polars
                    if operator == '==':
                        result_df = df.filter(pl.col(col_name) == value)
                    elif operator == '!=':
                        result_df = df.filter(pl.col(col_name) != value)
                    elif operator == '>':
                        result_df = df.filter(pl.col(col_name) > value)
                    elif operator == '<':
                        result_df = df.filter(pl.col(col_name) < value)
                    elif operator == '>=':
                        result_df = df.filter(pl.col(col_name) >= value)
                    elif operator == '<=':
                        result_df = df.filter(pl.col(col_name) <= value)
                    else:
                        return f"Opérateur '{operator}' non supporté dans Polars. Veuillez utiliser ==, !=, >, <, >= ou <=."
                else:
                    # Fallback: utiliser pandas pour la requête
                    result_df = pl.from_pandas(df_pd.query(query))

                # Générer la réponse
                header = f"Résultat de la requête '{query}' (traité avec Polars):\n"
                header += f"Trouvé {len(result_df)} lignes sur {len(df)} totales\n\n"

                if len(result_df) > 0:
                    # Limiter le nombre de lignes si trop grand
                    if len(result_df) > 1000:
                        display_df = result_df.head(1000)
                        header += f"Affichage limité aux 1000 premières lignes\n\n"
                    else:
                        display_df = result_df

                    return header + display_df.to_pandas().to_string(index=False)
                else:
                    return header + "Aucun résultat trouvé."

            except Exception as query_error:
                # Si la requête Polars échoue, essayer avec pandas
                result_df = df_pd.query(query)
                header = f"Résultat de la requête '{query}' (via pandas):\n"
                header += f"Trouvé {len(result_df)} lignes sur {len(df_pd)} totales\n\n"

                if len(result_df) > 0:
                    return header + result_df.to_string(index=False)
                else:
                    return header + "Aucun résultat trouvé."
        else:
            # Utiliser pandas pour les fichiers standards
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)

            # Exécuter la requête
            try:
                result_df = df.query(query)

                # Générer la réponse
                header = f"Résultat de la requête '{query}':\n"
                header += f"Trouvé {len(result_df)} lignes sur {len(df)} totales\n\n"

                if len(result_df) > 0:
                    return header + result_df.to_string(index=False)
                else:
                    return header + "Aucun résultat trouvé."

            except Exception as query_error:
                return f"Erreur dans la requête: {str(query_error)}"

    except Exception as e:
        return f"Erreur lors de l'exécution de la requête: {str(e)}"


@mcp.tool()
def create_bar_chart(file_path: str, x_column: str, y_column: str, sheet_name: str = None,
                     title: str = "Graphique à barres", max_points: int = 100,
                     aggregation: str = None) -> str:
    """
    Crée un graphique à barres à partir de données Excel.

    Args:
        file_path: Chemin du fichier Excel
        x_column: Nom de la colonne pour l'axe X
        y_column: Nom de la colonne pour l'axe Y
        sheet_name: Nom de la feuille (optionnel)
        title: Titre du graphique
        max_points: Nombre maximum de points à afficher (par défaut 100)
        aggregation: Méthode d'agrégation ('sum', 'mean', 'count', 'min', 'max')

    Returns:
        Image du graphique encodée en base64
    """
    try:
        # Gérer les chemins relatifs
        if not os.path.isabs(file_path):
            file_path = os.path.join(DEFAULT_EXCEL_DIR, file_path)

        # Vérifier que le fichier existe
        if not os.path.exists(file_path):
            return f"Erreur: Le fichier '{file_path}' n'existe pas."

        # Déterminer si le fichier est volumineux
        large_file = is_large_file(file_path)

        if large_file:
            # Utiliser Polars pour les fichiers volumineux
            if sheet_name:
                df_pd = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df_pd = pd.read_excel(file_path)

            # Convertir en DataFrame Polars
            df = pl.from_pandas(df_pd)

            # Vérifier que les colonnes existent
            if x_column not in df.columns:
                return f"Erreur: La colonne '{x_column}' n'existe pas dans le fichier."
            if y_column not in df.columns:
                return f"Erreur: La colonne '{y_column}' n'existe pas dans le fichier."

            # Appliquer l'agrégation si demandée
            if aggregation and aggregation.lower() in ['sum', 'mean', 'count', 'min', 'max']:
                agg_func = aggregation.lower()

                # Appliquer l'agrégation avec Polars
                if agg_func == 'sum':
                    df_agg = df.group_by(x_column).agg(pl.sum(y_column).alias(y_column))
                elif agg_func == 'mean':
                    df_agg = df.group_by(x_column).agg(pl.mean(y_column).alias(y_column))
                elif agg_func == 'count':
                    df_agg = df.group_by(x_column).agg(pl.count(y_column).alias(y_column))
                elif agg_func == 'min':
                    df_agg = df.group_by(x_column).agg(pl.min(y_column).alias(y_column))
                elif agg_func == 'max':
                    df_agg = df.group_by(x_column).agg(pl.max(y_column).alias(y_column))

                df = df_agg

            # Si toujours trop de lignes, échantillonner
            if len(df) > max_points:
                # Échantillonnage adaptatif
                if len(df) > 10 * max_points:
                    # Pour les très grands jeux de données, utiliser le quantile
                    quantiles = np.linspace(0, 1, max_points)
                    df_pd = df.to_pandas()  # Conversion à pandas pour l'échantillonnage quantile
                    indices = [int(q * (len(df_pd) - 1)) for q in quantiles]
                    df_pd = df_pd.iloc[indices].reset_index(drop=True)
                    chart_df = df_pd
                else:
                    # Pour les jeux de données moyennement grands, échantillonner régulièrement
                    step = max(1, len(df) // max_points)
                    df = df.slice(0, len(df), step)
                    chart_df = df.to_pandas()
            else:
                chart_df = df.to_pandas()

            # Créer le graphique
            plt.figure(figsize=(12, 7))
            plt.bar(chart_df[x_column], chart_df[y_column])

            # Ajout d'informations
            if aggregation:
                plt.title(f"{title} ({aggregation} par {x_column})")
            else:
                plt.title(title)

            plt.xlabel(x_column)
            plt.ylabel(y_column)
            plt.xticks(rotation=45)
            plt.grid(True, linestyle='--', alpha=0.7)
            plt.tight_layout()

            # Convertir le graphique en image base64
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=300)
            buffer.seek(0)
            image_base64 = base64.b64encode(buffer.read()).decode('utf-8')
            plt.close()

            return f"data:image/png;base64,{image_base64}"
        else:
            # Utiliser pandas pour les fichiers standards
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)

            # Vérifier que les colonnes existent
            if x_column not in df.columns:
                return f"Erreur: La colonne '{x_column}' n'existe pas dans le fichier."
            if y_column not in df.columns:
                return f"Erreur: La colonne '{y_column}' n'existe pas dans le fichier."

            # Appliquer l'agrégation si demandée
            if aggregation and aggregation.lower() in ['sum', 'mean', 'count', 'min', 'max']:
                agg_func = getattr(np, aggregation.lower())
                df = df.groupby(x_column).agg({y_column: agg_func}).reset_index()

            # Si toujours trop de lignes, échantillonner
            if len(df) > max_points:
                step = len(df) // max_points
                df = df.iloc[::step, :].reset_index(drop=True)

            # Créer le graphique
            plt.figure(figsize=(10, 6))
            plt.bar(df[x_column], df[y_column])
            plt.title(title)
            plt.xlabel(x_column)
            plt.ylabel(y_column)
            plt.xticks(rotation=45)
            plt.tight_layout()

            # Convertir le graphique en image base64
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png')
            buffer.seek(0)
            image_base64 = base64.b64encode(buffer.read()).decode('utf-8')
            plt.close()

            return f"data:image/png;base64,{image_base64}"

    except Exception as e:
        return f"Erreur lors de la création du graphique: {str(e)}"


@mcp.tool()
def create_pie_chart(file_path: str, labels_column: str, values_column: str, sheet_name: str = None,
                     title: str = "Graphique circulaire", max_segments: int = 10) -> str:
    """
    Crée un graphique circulaire à partir de données Excel.

    Args:
        file_path: Chemin du fichier Excel
        labels_column: Nom de la colonne pour les étiquettes
        values_column: Nom de la colonne pour les valeurs
        sheet_name: Nom de la feuille (optionnel)
        title: Titre du graphique
        max_segments: Nombre maximum de segments (regroupe les plus petits en "Autres")

    Returns:
        Description textuelle du graphique
    """
    try:
        # Gérer les chemins relatifs
        if not os.path.isabs(file_path):
            file_path = os.path.join(DEFAULT_EXCEL_DIR, file_path)

        # Déterminer si le fichier est volumineux
        large_file = is_large_file(file_path)

        if large_file:
            # Utiliser Polars pour les fichiers volumineux
            if sheet_name:
                df_pd = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df_pd = pd.read_excel(file_path)

            # Convertir en DataFrame Polars
            df = pl.from_pandas(df_pd)

            # Agréger les données
            aggregated_df = df.group_by(labels_column).agg(pl.sum(values_column).alias(values_column))

            # Trier par valeur décroissante pour identifier les segments principaux
            aggregated_df = aggregated_df.sort(values_column, descending=True)

            # Si trop de segments, regrouper les plus petits en "Autres"
            if len(aggregated_df) > max_segments:
                main_segments = aggregated_df.head(max_segments - 1)
                other_segments = aggregated_df.slice(max_segments - 1, len(aggregated_df))

                # Calculer la somme des segments "Autres"
                other_sum = other_segments[values_column].sum()

                # Créer un nouveau DataFrame avec les segments principaux et "Autres"
                main_df = main_segments.to_pandas()
                other_row = pd.DataFrame({labels_column: ["Autres"], values_column: [other_sum]})
                chart_df = pd.concat([main_df, other_row], ignore_index=True)
            else:
                chart_df = aggregated_df.to_pandas()

            # Calculer les pourcentages
            total = chart_df[values_column].sum()
            chart_df['percentage'] = chart_df[values_column] / total * 100

            # Créer une description textuelle du graphique
            result = f"Graphique circulaire: {title}\n\n"
            result += f"Distribution de {values_column} par {labels_column}:\n"

            for i, row in chart_df.iterrows():
                label = row[labels_column]
                value = row[values_column]
                percentage = row['percentage']
                result += f"- {label}: {value:,.2f} ({percentage:.1f}%)\n"

            return result
        else:
            # Utiliser pandas pour les fichiers standards
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)

            # Agréger les données si nécessaire
            aggregated_df = df.groupby(labels_column)[values_column].sum().reset_index()

            # Calculer les pourcentages
            total = aggregated_df[values_column].sum()
            percentages = [(value / total * 100) for value in aggregated_df[values_column]]

            # Créer une description textuelle du graphique
            result = f"Graphique circulaire: {title}\n\n"
            result += f"Distribution de {values_column} par {labels_column}:\n"

            for i, (label, value) in enumerate(zip(aggregated_df[labels_column], aggregated_df[values_column])):
                result += f"- {label}: {value:,.2f} ({percentages[i]:.1f}%)\n"

            return result

    except Exception as e:
        return f"Erreur lors de la création du graphique: {str(e)}"


@mcp.tool()
def create_line_chart(file_path: str, x_column: str, y_columns: str, sheet_name: str = None,
                      title: str = "Graphique linéaire", max_points: int = 500,
                      aggregation: str = None) -> str:
    """
    Crée un graphique linéaire à partir de données Excel.

    Args:
        file_path: Chemin du fichier Excel
        x_column: Nom de la colonne pour l'axe X
        y_columns: Noms des colonnes pour l'axe Y (séparés par des virgules)
        sheet_name: Nom de la feuille (optionnel)
        title: Titre du graphique
        max_points: Nombre maximum de points par série (par défaut 500)
        aggregation: Méthode d'agrégation ('sum', 'mean', 'count', 'min', 'max')

    Returns:
        Image du graphique encodée en base64
    """
    try:
        # Gérer les chemins relatifs
        if not os.path.isabs(file_path):
            file_path = os.path.join(DEFAULT_EXCEL_DIR, file_path)

        # Vérifier que le fichier existe
        if not os.path.exists(file_path):
            return f"Erreur: Le fichier '{file_path}' n'existe pas."

        # Déterminer si le fichier est volumineux
        large_file = is_large_file(file_path)

        # Traiter les colonnes Y
        y_columns_list = [col.strip() for col in y_columns.split(',')]

        if large_file:
            # Utiliser Polars pour les fichiers volumineux
            if sheet_name:
                df_pd = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df_pd = pd.read_excel(file_path)

            # Convertir en DataFrame Polars
            df = pl.from_pandas(df_pd)

            # Vérifier que la colonne X existe
            if x_column not in df.columns:
                return f"Erreur: La colonne '{x_column}' n'existe pas dans le fichier."

            # Vérifier que les colonnes Y existent
            for col in y_columns_list:
                if col not in df.columns:
                    return f"Erreur: La colonne '{col}' n'existe pas dans le fichier."

            # Appliquer l'agrégation si demandée
            if aggregation and aggregation.lower() in ['sum', 'mean', 'count', 'min', 'max']:
                agg_func = aggregation.lower()

                # Préparer les expressions d'agrégation pour chaque colonne Y
                agg_exprs = []
                for col in y_columns_list:
                    if agg_func == 'sum':
                        agg_exprs.append(pl.sum(col).alias(col))
                    elif agg_func == 'mean':
                        agg_exprs.append(pl.mean(col).alias(col))
                    elif agg_func == 'count':
                        agg_exprs.append(pl.count(col).alias(col))
                    elif agg_func == 'min':
                        agg_exprs.append(pl.min(col).alias(col))
                    elif agg_func == 'max':
                        agg_exprs.append(pl.max(col).alias(col))

                # Appliquer l'agrégation
                df = df.group_by(x_column).agg(agg_exprs)

            # Si toujours trop de points, échantillonner
            if len(df) > max_points:
                # Échantillonnage adaptatif selon la taille des données
                if len(df) > 10 * max_points:
                    # Pour les très grands jeux de données, utiliser le quantile
                    df_pd = df.to_pandas()  # Conversion à pandas pour l'échantillonnage quantile
                    quantiles = np.linspace(0, 1, max_points)
                    indices = [int(q * (len(df_pd) - 1)) for q in quantiles]
                    chart_df = df_pd.iloc[indices].reset_index(drop=True)
                else:
                    # Pour les jeux de données moyennement grands, échantillonner régulièrement
                    step = max(1, len(df) // max_points)
                    df = df.slice(0, len(df), step)
                    chart_df = df.to_pandas()
            else:
                chart_df = df.to_pandas()

            # Trier les données par la colonne X
            chart_df = chart_df.sort_values(by=x_column)

            # Créer le graphique
            plt.figure(figsize=(14, 8))

            for col in y_columns_list:
                plt.plot(chart_df[x_column], chart_df[col], marker='o', markersize=3, linewidth=2, alpha=0.7, label=col)

            # Ajout d'informations
            if aggregation:
                plt.title(f"{title} ({aggregation} par {x_column})")
            else:
                plt.title(title)

            plt.xlabel(x_column)
            plt.ylabel("Valeurs")
            plt.legend(loc='best')
            plt.grid(True, linestyle='--', alpha=0.7)

            # Rendre l'axe des x plus lisible
            if len(chart_df) > 30:
                # Si beaucoup de points, n'afficher qu'un sous-ensemble des ticks
                tick_step = max(1, len(chart_df) // 20)
                plt.xticks(chart_df[x_column][::tick_step], rotation=45)
            else:
                plt.xticks(rotation=45)

            plt.tight_layout()

            # Convertir le graphique en image base64
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=300)
            buffer.seek(0)
            image_base64 = base64.b64encode(buffer.read()).decode('utf-8')
            plt.close()

            return f"data:image/png;base64,{image_base64}"
        else:
            # Utiliser pandas pour les fichiers standards
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)

            # Vérifier que la colonne X existe
            if x_column not in df.columns:
                return f"Erreur: La colonne '{x_column}' n'existe pas dans le fichier."

            # Vérifier que les colonnes Y existent
            for col in y_columns_list:
                if col not in df.columns:
                    return f"Erreur: La colonne '{col}' n'existe pas dans le fichier."

            # Appliquer l'agrégation si demandée
            if aggregation and aggregation.lower() in ['sum', 'mean', 'count', 'min', 'max']:
                agg_dict = {col: aggregation.lower() for col in y_columns_list}
                df = df.groupby(x_column).agg(agg_dict).reset_index()

            # Si toujours trop de lignes, échantillonner
            if len(df) > max_points:
                step = len(df) // max_points
                df = df.iloc[::step, :].reset_index(drop=True)

            # Trier par la colonne X pour un graphique plus cohérent
            df = df.sort_values(by=x_column)

            # Créer le graphique
            plt.figure(figsize=(12, 7))
            for col in y_columns_list:
                plt.plot(df[x_column], df[col], marker='o', label=col)

            plt.title(title)
            plt.xlabel(x_column)
            plt.ylabel("Valeurs")
            plt.legend()
            plt.grid(True, linestyle='--', alpha=0.7)
            plt.xticks(rotation=45)
            plt.tight_layout()

            # Convertir le graphique en image base64
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png')
            buffer.seek(0)
            image_base64 = base64.b64encode(buffer.read()).decode('utf-8')
            plt.close()

            return f"data:image/png;base64,{image_base64}"

    except Exception as e:
        return f"Erreur lors de la création du graphique: {str(e)}"


@mcp.tool()
def create_scatter_plot(file_path: str, x_column: str, y_column: str, color_column: str = None, sheet_name: str = None,
                        title: str = "Nuage de points", max_points: int = 1000) -> str:
    """
    Crée un nuage de points à partir de données Excel.

    Args:
        file_path: Chemin du fichier Excel
        x_column: Nom de la colonne pour l'axe X
        y_column: Nom de la colonne pour l'axe Y
        color_column: Nom de la colonne pour la couleur des points (optionnel)
        sheet_name: Nom de la feuille (optionnel)
        title: Titre du graphique
        max_points: Nombre maximum de points à afficher (échantillonnage si dépassé)

    Returns:
        Image du graphique encodée en base64
    """
    try:
        # Gérer les chemins relatifs
        if not os.path.isabs(file_path):
            file_path = os.path.join(DEFAULT_EXCEL_DIR, file_path)

        # Vérifier que le fichier existe
        if not os.path.exists(file_path):
            return f"Erreur: Le fichier '{file_path}' n'existe pas."

        # Déterminer si le fichier est volumineux
        large_file = is_large_file(file_path)

        if large_file:
            # Utiliser Polars pour les fichiers volumineux
            if sheet_name:
                df_pd = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df_pd = pd.read_excel(file_path)

            # Convertir en DataFrame Polars
            df = pl.from_pandas(df_pd)

            # Vérifier que les colonnes existent
            if x_column not in df.columns:
                return f"Erreur: La colonne '{x_column}' n'existe pas dans le fichier."
            if y_column not in df.columns:
                return f"Erreur: La colonne '{y_column}' n'existe pas dans le fichier."
            if color_column and color_column not in df.columns:
                return f"Erreur: La colonne '{color_column}' n'existe pas dans le fichier."

            # Si trop de points, échantillonner
            if len(df) > max_points:
                # Échantillonnage stratifié si une colonne couleur est utilisée
                if color_column:
                    df_pd = df.to_pandas()

                    # Stratifier par la colonne couleur
                    unique_colors = df_pd[color_column].unique()
                    samples = []

                    for color in unique_colors:
                        subset = df_pd[df_pd[color_column] == color]
                        # Déterminer combien de points prendre de chaque strate
                        n_samples = max(1, int(max_points * len(subset) / len(df_pd)))
                        if len(subset) > n_samples:
                            samples.append(subset.sample(n=n_samples))
                        else:
                            samples.append(subset)

                    # Combiner les échantillons
                    chart_df = pd.concat(samples, ignore_index=True)
                else:
                    # Échantillonnage aléatoire simple
                    df_pd = df.to_pandas()
                    chart_df = df_pd.sample(n=max_points)
            else:
                chart_df = df.to_pandas()

            # Créer le graphique
            plt.figure(figsize=(12, 9))

            if color_column:
                scatter = plt.scatter(chart_df[x_column], chart_df[y_column],
                                      c=chart_df[color_column], cmap='viridis',
                                      alpha=0.7, s=30, edgecolors='gray')
                plt.colorbar(scatter, label=color_column)
            else:
                plt.scatter(chart_df[x_column], chart_df[y_column],
                            alpha=0.7, s=30, edgecolors='gray')

            plt.title(title)
            plt.xlabel(x_column)
            plt.ylabel(y_column)
            plt.grid(True, linestyle='--', alpha=0.3)
            plt.tight_layout()

            # Convertir le graphique en image base64
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=300)
            buffer.seek(0)
            image_base64 = base64.b64encode(buffer.read()).decode('utf-8')
            plt.close()

            return f"data:image/png;base64,{image_base64}"
        else:
            # Utiliser pandas pour les fichiers standards
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)

            # Vérifier que les colonnes existent
            if x_column not in df.columns:
                return f"Erreur: La colonne '{x_column}' n'existe pas dans le fichier."
            if y_column not in df.columns:
                return f"Erreur: La colonne '{y_column}' n'existe pas dans le fichier."
            if color_column and color_column not in df.columns:
                return f"Erreur: La colonne '{color_column}' n'existe pas dans le fichier."

            # Si trop de points, échantillonner
            if len(df) > max_points:
                df = df.sample(n=max_points)

            # Créer le graphique
            plt.figure(figsize=(10, 8))

            if color_column:
                scatter = plt.scatter(df[x_column], df[y_column], c=df[color_column], cmap='viridis', alpha=0.6)
                plt.colorbar(scatter, label=color_column)
            else:
                plt.scatter(df[x_column], df[y_column], alpha=0.6)

            plt.title(title)
            plt.xlabel(x_column)
            plt.ylabel(y_column)
            plt.grid(True, linestyle='--', alpha=0.3)
            plt.tight_layout()

            # Convertir le graphique en image base64
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png')
            buffer.seek(0)
            image_base64 = base64.b64encode(buffer.read()).decode('utf-8')
            plt.close()

            return f"data:image/png;base64,{image_base64}"

    except Exception as e:
        return f"Erreur lors de la création du graphique: {str(e)}"


@mcp.tool()
def get_column_names(file_path: str, sheet_name: str = None) -> str:
    """
    Retourne la liste des noms de colonnes d'un fichier Excel.

    Args:
        file_path: Chemin du fichier Excel
        sheet_name: Nom de la feuille (optionnel)

    Returns:
        Liste des noms de colonnes
    """
    try:
        # Gérer les chemins relatifs
        if not os.path.isabs(file_path):
            file_path = os.path.join(DEFAULT_EXCEL_DIR, file_path)

        # Vérifier que le fichier existe
        if not os.path.exists(file_path):
            return f"Erreur: Le fichier '{file_path}' n'existe pas."

        # Déterminer si le fichier est volumineux
        large_file = is_large_file(file_path)

        if large_file:
            # Pour les noms de colonnes, on peut utiliser la technique d'échantillonnage
            # pour éviter de charger tout le fichier
            # Utiliser pandas avec nrows limité
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)
            else:
                df = pd.read_excel(file_path, nrows=5)

            # Obtenir et formater les noms de colonnes
            columns = df.columns.tolist()
            result = f"Colonnes dans '{os.path.basename(file_path)}'"
            if sheet_name:
                result += f" (feuille '{sheet_name}')"
            result += ":\n\n"

            for i, col in enumerate(columns, 1):
                result += f"{i}. {col}\n"

            result += f"\n(Fichier volumineux détecté: {os.path.getsize(file_path) / (1024 * 1024):.1f} Mo)"

            return result
        else:
            # Utiliser pandas pour les fichiers standards
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)

            # Obtenir et formater les noms de colonnes
            columns = df.columns.tolist()
            result = f"Colonnes dans '{os.path.basename(file_path)}'"
            if sheet_name:
                result += f" (feuille '{sheet_name}')"
            result += ":\n\n"

            for i, col in enumerate(columns, 1):
                result += f"{i}. {col}\n"

            return result

    except Exception as e:
        return f"Erreur lors de la récupération des noms de colonnes: {str(e)}"


@mcp.tool()
def create_excel_chart(file_path: str, x_column: str, y_columns: str, chart_type: str = "line",
                       sheet_name: str = None, new_sheet_name: str = None, title: str = "Graphique",
                       aggregation: str = None) -> str:
    """
    Crée un graphique et l'ajoute directement dans une nouvelle feuille du fichier Excel.

    Args:
        file_path: Chemin du fichier Excel
        x_column: Nom de la colonne pour l'axe X
        y_columns: Noms des colonnes pour l'axe Y (séparés par des virgules)
        chart_type: Type de graphique (line, bar, pie, scatter)
        sheet_name: Nom de la feuille source des données (si None, utilise la première feuille)
        new_sheet_name: Nom de la nouvelle feuille à créer (si None, utilise "Graphique_[timestamp]")
        title: Titre du graphique
        aggregation: Méthode d'agrégation ('sum', 'mean', 'count', 'min', 'max')

    Returns:
        Message de confirmation ou d'erreur
    """
    try:
        # Gérer les chemins relatifs
        if not os.path.isabs(file_path):
            file_path = os.path.join(DEFAULT_EXCEL_DIR, file_path)

        # Vérifier que le fichier existe
        if not os.path.exists(file_path):
            return f"Erreur: Le fichier '{file_path}' n'existe pas."

        # Déterminer si le fichier est volumineux
        large_file = is_large_file(file_path)

        # Importer les bibliothèques nécessaires
        from openpyxl import load_workbook
        from openpyxl.chart import (
            LineChart, BarChart, PieChart, ScatterChart,
            Reference, Series
        )
        from openpyxl.utils.dataframe import dataframe_to_rows
        import datetime

        # Traiter les colonnes Y
        y_columns_list = [col.strip() for col in y_columns.split(',')]

        if large_file:
            # Pour les fichiers volumineux, on charge avec Polars puis on extrait un échantillon agrégé
            if sheet_name:
                df_pd = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df_pd = pd.read_excel(file_path)

            # Convertir en DataFrame Polars
            df_pl = pl.from_pandas(df_pd)

            # Vérifier que les colonnes existent
            if x_column not in df_pl.columns:
                return f"Erreur: La colonne '{x_column}' n'existe pas dans le fichier."

            for col in y_columns_list:
                if col not in df_pl.columns:
                    return f"Erreur: La colonne '{col}' n'existe pas dans le fichier."

            # Appliquer l'agrégation si demandée
            if aggregation and aggregation.lower() in ['sum', 'mean', 'count', 'min', 'max']:
                agg_func = aggregation.lower()

                # Préparer les expressions d'agrégation pour chaque colonne Y
                agg_exprs = []
                for col in y_columns_list:
                    if agg_func == 'sum':
                        agg_exprs.append(pl.sum(col).alias(col))
                    elif agg_func == 'mean':
                        agg_exprs.append(pl.mean(col).alias(col))
                    elif agg_func == 'count':
                        agg_exprs.append(pl.count(col).alias(col))
                    elif agg_func == 'min':
                        agg_exprs.append(pl.min(col).alias(col))
                    elif agg_func == 'max':
                        agg_exprs.append(pl.max(col).alias(col))

                # Appliquer l'agrégation
                df_pl = df_pl.group_by(x_column).agg(agg_exprs)

            # Limiter le nombre de points si nécessaire
            if len(df_pl) > 1000:
                # Échantillonnage adaptatif
                step = max(1, len(df_pl) // 1000)
                df_pl = df_pl.slice(0, len(df_pl), step)

            # Trier par la colonne X
            df_pl = df_pl.sort(x_column)

            # Convertir le résultat final en pandas pour l'export
            chart_data = df_pl.to_pandas()
        else:
            # Utiliser pandas pour les fichiers standards
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)

            # Vérifier que les colonnes existent
            if x_column not in df.columns:
                return f"Erreur: La colonne '{x_column}' n'existe pas dans le fichier."

            for col in y_columns_list:
                if col not in df.columns:
                    return f"Erreur: La colonne '{col}' n'existe pas dans le fichier."

            # Appliquer l'agrégation si demandée
            if aggregation and aggregation.lower() in ['sum', 'mean', 'count', 'min', 'max']:
                agg_dict = {col: aggregation.lower() for col in y_columns_list}
                df = df.groupby(x_column).agg(agg_dict).reset_index()

            # Extraire les données pertinentes
            chart_data = df[[x_column] + y_columns_list].copy()

        # Créer un nom pour la nouvelle feuille si non spécifié
        if not new_sheet_name:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            new_sheet_name = f"Graphique_{timestamp}"

        # Charger le classeur avec openpyxl
        wb = load_workbook(file_path)

        # Créer une nouvelle feuille
        if new_sheet_name in wb.sheetnames:
            # Si la feuille existe déjà, on la supprime et on la recrée
            idx = wb.sheetnames.index(new_sheet_name)
            wb.remove(wb.worksheets[idx])

        ws = wb.create_sheet(title=new_sheet_name)

        # Écrire les données dans la nouvelle feuille
        for r in dataframe_to_rows(chart_data, index=False, header=True):
            ws.append(r)

        # Créer le graphique selon le type demandé
        if chart_type.lower() == "line":
            chart = LineChart()
        elif chart_type.lower() == "bar":
            chart = BarChart()
        elif chart_type.lower() == "pie":
            if len(y_columns_list) > 1:
                return "Erreur: Le graphique circulaire ne peut utiliser qu'une seule colonne Y."
            chart = PieChart()
        elif chart_type.lower() == "scatter":
            chart = ScatterChart()
        else:
            return f"Erreur: Type de graphique '{chart_type}' non supporté."

        # Configurer le graphique
        chart_title = title
        if aggregation:
            chart_title += f" ({aggregation})"
        chart.title = chart_title
        chart.x_axis.title = x_column
        chart.y_axis.title = "Valeurs" if len(y_columns_list) > 1 else y_columns_list[0]

        # Définir les plages de données
        data = Reference(ws, min_col=2, min_row=1, max_row=len(chart_data) + 1, max_col=len(y_columns_list) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(chart_data) + 1)

        # Ajouter les données au graphique
        chart.add_data(data, titles_from_data=True)

        # Pour les graphiques autres que circulaires, définir les catégories
        if chart_type.lower() != "pie":
            chart.set_categories(cats)

        # Ajouter le graphique à la feuille
        ws.add_chart(chart, "A" + str(len(chart_data) + 5))

        # Enregistrer le fichier
        wb.save(file_path)

        # Message de confirmation
        message = f"Graphique '{title}' créé avec succès dans la feuille '{new_sheet_name}' du fichier '{os.path.basename(file_path)}'."
        if large_file:
            message += f"\n(Fichier volumineux traité avec Polars pour une performance optimale)"
        if aggregation:
            message += f"\n(Données agrégées par '{aggregation}' pour une meilleure performance)"

        return message

    except Exception as e:
        return f"Erreur lors de la création du graphique dans Excel: {str(e)}"


@mcp.tool()
def read_excel_optimized(file_path: str, sheet_name: str = None,
                         chunk_size: int = 1000, columns: str = None) -> str:
    """
    Lit un fichier Excel volumineux de manière optimisée.

    Args:
        file_path: Chemin du fichier Excel
        sheet_name: Nom de la feuille à lire (si None, lit la première feuille)
        chunk_size: Taille des morceaux à lire (nombre de lignes)
        columns: Colonnes à lire (séparées par virgules, si None, lit toutes les colonnes)

    Returns:
        Résumé des données lues
    """
    try:
        # Gérer les chemins relatifs
        if not os.path.isabs(file_path):
            file_path = os.path.join(DEFAULT_EXCEL_DIR, file_path)

        # Vérifier que le fichier existe
        if not os.path.exists(file_path):
            return f"Erreur: Le fichier '{file_path}' n'existe pas."

        # Préparer les colonnes à lire
        cols_to_use = None
        if columns:
            cols_to_use = [col.strip() for col in columns.split(',')]

        # Vérifier si le fichier est volumineux
        file_size_mb = os.path.getsize(file_path) / (1024 * 1024)

        if file_size_mb > LARGE_FILE_THRESHOLD:
            # Pour les fichiers Excel volumineux, on utilise une approche optimisée
            # Lire seulement les premières lignes pour obtenir les métadonnées
            if sheet_name:
                meta_df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5, usecols=cols_to_use)
            else:
                meta_df = pd.read_excel(file_path, nrows=5, usecols=cols_to_use)

            # Obtenir le nombre total de lignes (approximation pour les fichiers Excel)
            # Cette approche peut être lente pour les très grands fichiers
            # Dans ce cas, on fait une estimation basée sur la taille du fichier
            try:
                if file_size_mb > 500:  # Si > 500 Mo, on estime
                    avg_row_size = 500  # octets par ligne (approximation)
                    est_rows = int((file_size_mb * 1024 * 1024) / avg_row_size)
                    total_rows = f"~{est_rows:,} (estimation)"
                else:
                    # Tenter de compter les lignes réelles (peut être lent)
                    if sheet_name:
                        full_df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=[0])
                    else:
                        full_df = pd.read_excel(file_path, usecols=[0])
                    total_rows = f"{len(full_df):,}"
            except:
                total_rows = "Non déterminé (fichier très volumineux)"

            # Lire un échantillon pour l'affichage
            if sheet_name:
                sample_df = pd.read_excel(file_path, sheet_name=sheet_name,
                                          nrows=min(chunk_size, 1000),
                                          usecols=cols_to_use)
            else:
                sample_df = pd.read_excel(file_path, nrows=min(chunk_size, 1000),
                                          usecols=cols_to_use)

            # Convertir en Polars pour des statistiques rapides
            sample_pl = pl.from_pandas(sample_df)

            # Générer un résumé
            result = f"Fichier: {os.path.basename(file_path)} ({file_size_mb:.1f} Mo)\n"
            result += f"Feuille: {sheet_name or 'par défaut'}\n"
            result += f"Nombre total de lignes: {total_rows}\n"
            result += f"Nombre de colonnes: {len(sample_pl.columns)}\n"
            if columns:
                result += f"Colonnes sélectionnées: {columns}\n"
            result += "\nAperçu des données (premières lignes):\n"
            result += sample_df.head(10).to_string(index=False)

            # Ajouter des statistiques de base sur l'échantillon
            result += "\n\nStatistiques sur l'échantillon:\n"

            # Identifier les colonnes numériques
            num_cols = []
            for col in sample_pl.columns:
                if sample_pl[col].dtype in [pl.Float32, pl.Float64, pl.Int32, pl.Int64]:
                    num_cols.append(col)

            if num_cols:
                result += "\nStatistiques pour colonnes numériques:\n"
                for col in num_cols[:5]:  # Limiter à 5 colonnes pour la lisibilité
                    try:
                        result += f"- {col}: min={sample_pl[col].min()}, max={sample_pl[col].max()}, "
                        result += f"moyenne={sample_pl[col].mean():.2f}, médiane={sample_pl[col].median()}\n"
                    except:
                        result += f"- {col}: statistiques non disponibles\n"

                if len(num_cols) > 5:
                    result += f"... et {len(num_cols) - 5} autres colonnes numériques\n"

            # Message pour l'utilisateur
            result += "\nNote: Ce fichier est volumineux. Pour de meilleures performances:\n"
            result += "1. Utilisez excel_query pour filtrer les données avant traitement\n"
            result += "2. Utilisez des fonctions d'agrégation dans les graphiques\n"
            result += "3. Pour les graphiques, ajoutez le paramètre 'aggregation' (sum, mean, etc.)\n"

            return result
        else:
            # Pour les fichiers de taille normale, on utilise pandas standard
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=cols_to_use)
            else:
                df = pd.read_excel(file_path, usecols=cols_to_use)

            # Générer un résumé
            result = f"Fichier: {os.path.basename(file_path)}\n"
            result += f"Feuille: {sheet_name or 'par défaut'}\n"
            result += f"Dimensions: {df.shape[0]:,} lignes × {df.shape[1]} colonnes\n"
            if columns:
                result += f"Colonnes sélectionnées: {columns}\n"
            result += "\nAperçu des données:\n"
            result += df.head(20).to_string(index=False)

            # Ajouter des statistiques de base
            result += "\n\nStatistiques de base:\n"
            result += df.describe().to_string()

            return result

    except Exception as e:
        return f"Erreur lors de la lecture optimisée: {str(e)}"


@mcp.tool()
def create_aggregated_chart(file_path: str, x_column: str, y_column: str,
                            aggregation: str = "sum", chart_type: str = "bar",
                            sheet_name: str = None, title: str = None,
                            limit: int = 20) -> str:
    """
    Crée un graphique à partir de données agrégées pour gérer les fichiers volumineux.

    Args:
        file_path: Chemin du fichier Excel
        x_column: Nom de la colonne pour regrouper les données (axe X)
        y_column: Nom de la colonne à agréger (axe Y)
        aggregation: Méthode d'agrégation ('sum', 'mean', 'count', 'min', 'max')
        chart_type: Type de graphique ('bar', 'line', 'pie')
        sheet_name: Nom de la feuille (optionnel)
        title: Titre du graphique (optionnel)
        limit: Limite du nombre de groupes à afficher (top N)

    Returns:
        Image du graphique encodée en base64
    """
    try:
        # Gérer les chemins relatifs
        if not os.path.isabs(file_path):
            file_path = os.path.join(DEFAULT_EXCEL_DIR, file_path)

        # Vérifier que le fichier existe
        if not os.path.exists(file_path):
            return f"Erreur: Le fichier '{file_path}' n'existe pas."

        # Déterminer si le fichier est volumineux
        large_file = is_large_file(file_path)

        # Préparer le titre
        if not title:
            title = f"{aggregation.capitalize()} de {y_column} par {x_column}"

        if large_file:
            # Utiliser Polars pour les fichiers volumineux
            if sheet_name:
                df_pd = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df_pd = pd.read_excel(file_path)

            # Convertir en DataFrame Polars
            df = pl.from_pandas(df_pd)

            # Vérifier que les colonnes existent
            if x_column not in df.columns:
                return f"Erreur: La colonne '{x_column}' n'existe pas dans le fichier."
            if y_column not in df.columns:
                return f"Erreur: La colonne '{y_column}' n'existe pas dans le fichier."

            # Appliquer l'agrégation
            if aggregation.lower() == 'sum':
                df_agg = df.group_by(x_column).agg(pl.sum(y_column).alias(y_column))
            elif aggregation.lower() == 'mean':
                df_agg = df.group_by(x_column).agg(pl.mean(y_column).alias(y_column))
            elif aggregation.lower() == 'count':
                df_agg = df.group_by(x_column).agg(pl.count(y_column).alias(y_column))
            elif aggregation.lower() == 'min':
                df_agg = df.group_by(x_column).agg(pl.min(y_column).alias(y_column))
            elif aggregation.lower() == 'max':
                df_agg = df.group_by(x_column).agg(pl.max(y_column).alias(y_column))
            else:
                return f"Erreur: Méthode d'agrégation '{aggregation}' non supportée. Utilisez 'sum', 'mean', 'count', 'min' ou 'max'."

            # Trier et limiter le nombre de groupes
            df_agg = df_agg.sort(y_column, descending=True).head(limit)

            # Convertir en pandas pour le graphique
            chart_df = df_agg.to_pandas()
        else:
            # Utiliser pandas pour les fichiers standards
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)

            # Vérifier que les colonnes existent
            if x_column not in df.columns:
                return f"Erreur: La colonne '{x_column}' n'existe pas dans le fichier."
            if y_column not in df.columns:
                return f"Erreur: La colonne '{y_column}' n'existe pas dans le fichier."

            # Appliquer l'agrégation
            df_agg = df.groupby(x_column).agg({y_column: aggregation.lower()}).reset_index()

            # Trier et limiter le nombre de groupes
            df_agg = df_agg.sort_values(by=y_column, ascending=False).head(limit)

            chart_df = df_agg

        # Créer le graphique selon le type demandé
        plt.figure(figsize=(12, 8))

        if chart_type.lower() == 'bar':
            plt.bar(chart_df[x_column], chart_df[y_column])
        elif chart_type.lower() == 'line':
            plt.plot(chart_df[x_column], chart_df[y_column], marker='o', linewidth=2)
        elif chart_type.lower() == 'pie':
            # Limiter à 10 segments pour les graphiques circulaires
            if len(chart_df) > 10:
                pie_df = chart_df.head(9)
                other_sum = chart_df.iloc[9:][y_column].sum()
                other_df = pd.DataFrame({x_column: ['Autres'], y_column: [other_sum]})
                pie_df = pd.concat([pie_df, other_df])
                plt.pie(pie_df[y_column], labels=pie_df[x_column], autopct='%1.1f%%')
            else:
                plt.pie(chart_df[y_column], labels=chart_df[x_column], autopct='%1.1f%%')
        else:
            return f"Erreur: Type de graphique '{chart_type}' non supporté. Utilisez 'bar', 'line' ou 'pie'."

        plt.title(title)
        if chart_type.lower() != 'pie':
            plt.xlabel(x_column)
            plt.ylabel(f"{aggregation.capitalize()} de {y_column}")
            plt.grid(True, linestyle='--', alpha=0.7)
            plt.xticks(rotation=45)

        plt.tight_layout()

        # Convertir le graphique en image base64
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=300)
        buffer.seek(0)
        image_base64 = base64.b64encode(buffer.read()).decode('utf-8')
        plt.close()

        return f"data:image/png;base64,{image_base64}"

    except Exception as e:
        return f"Erreur lors de la création du graphique agrégé: {str(e)}"


@mcp.resource("excel://{file_path}")
def excel_resource(file_path: str) -> str:
    """
    Ressource pour accéder aux fichiers Excel.
    Exemple d'accès: excel://mon_fichier.xlsx
    """
    try:
        # Gérer les chemins relatifs
        if not os.path.isabs(file_path):
            file_path = os.path.join(DEFAULT_EXCEL_DIR, file_path)

        # Vérifier que le fichier existe
        if not os.path.exists(file_path):
            return f"Erreur: Le fichier '{file_path}' n'existe pas."

        # Déterminer si le fichier est volumineux
        large_file = is_large_file(file_path)

        # Obtenir la structure du fichier
        result = {}

        # Utiliser pandas ou polars selon la taille du fichier
        if large_file:
            # Pour les fichiers volumineux, on ne charge pas tout le contenu
            excel_file = pd.ExcelFile(file_path)

            # Obtenir les informations sur toutes les feuilles
            result["file_name"] = os.path.basename(file_path)
            result["sheets"] = excel_file.sheet_names
            result["file_size"] = f"{os.path.getsize(file_path) / (1024 * 1024):.2f} MB"
            result["last_modified"] = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
            result["large_file"] = True

            # Obtenir des informations de base sur chaque feuille
            sheet_info = {}
            for sheet in excel_file.sheet_names:
                # Lire seulement les premières lignes pour obtenir les colonnes
                df_sample = pd.read_excel(file_path, sheet_name=sheet, nrows=5)

                # Estimer le nombre de lignes total (peut être lent)
                try:
                    if os.path.getsize(file_path) > 500 * 1024 * 1024:  # > 500 Mo
                        rows_est = "Trop volumineux pour estimation rapide"
                    else:
                        df_count = pd.read_excel(file_path, sheet_name=sheet, usecols=[0])
                        rows_est = len(df_count)
                except:
                    rows_est = "Non déterminé"

                sheet_info[sheet] = {
                    "rows": rows_est,
                    "columns": len(df_sample.columns),
                    "column_names": df_sample.columns.tolist()
                }

            result["sheet_details"] = sheet_info
            result["optimized"] = "Analyse optimisée pour fichier volumineux avec Polars"
        else:
            # Pour les fichiers standards, on utilise pandas
            excel_file = pd.ExcelFile(file_path)

            # Obtenir les informations sur toutes les feuilles
            result["file_name"] = os.path.basename(file_path)
            result["sheets"] = excel_file.sheet_names
            result["file_size"] = f"{os.path.getsize(file_path) / 1024:.2f} KB"
            result["last_modified"] = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')

            # Obtenir des informations de base sur chaque feuille
            sheet_info = {}
            for sheet in excel_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet)
                sheet_info[sheet] = {
                    "rows": len(df),
                    "columns": len(df.columns),
                    "column_names": df.columns.tolist()
                }

            result["sheet_details"] = sheet_info

        # Formater le résultat en JSON pour lisibilité
        return json.dumps(result, indent=2)

    except Exception as e:
        return f"Erreur lors de l'accès au fichier Excel: {str(e)}"


if __name__ == "__main__":
    mcp.run()